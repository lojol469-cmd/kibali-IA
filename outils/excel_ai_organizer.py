"""
Outil d'organisation intelligente de données Excel par IA
Organise automatiquement tous types de données : coordonnées GPS, tableaux, classifications
avec précision ultra-dynamique et structuration intelligente
"""

import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import re
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional
import json

class ExcelAIOrganizer:
    """
    Organisateur intelligent de fichiers Excel avec IA
    - Détecte automatiquement le type de données
    - Organise selon le contexte (géospatial, invités, finances, scientifique, etc.)
    - Classifications ultra-précises et dynamiques
    - Support coordonnées GPS, timestamps, données numériques, textuelles
    """
    
    def __init__(self):
        self.data_types = {
            'coordinates': r'([-+]?\d+\.\d+)[,\s]+' + r'([-+]?\d+\.\d+)',
            'latitude': r'([-+]?\d+\.\d+)[\s]*[NS]?',
            'longitude': r'([-+]?\d+\.\d+)[\s]*[EW]?',
            'phone': r'\+?\d{1,4}[\s\-]?\(?\d{1,4}\)?[\s\-]?\d{1,4}[\s\-]?\d{1,9}',
            'email': r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
            'date': r'\d{1,4}[-/\.]\d{1,2}[-/\.]\d{1,4}',
            'time': r'\d{1,2}:\d{2}(:\d{2})?',
            'currency': r'[$€£¥]\s*[\d,]+\.?\d*|[\d,]+\.?\d*\s*[$€£¥]',
            'percentage': r'\d+\.?\d*\s*%',
            'url': r'https?://[^\s<>"{}|\\^[\]`]+',
            'ip_address': r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}',
        }
        
        self.color_schemes = {
            'geographic': {'header': '4472C4', 'accent': '70AD47', 'highlight': 'FFC000'},
            'financial': {'header': '00B050', 'accent': '92D050', 'highlight': 'FFFF00'},
            'scientific': {'header': '7030A0', 'accent': 'C65911', 'highlight': 'FF6600'},
            'events': {'header': 'E91E63', 'accent': '9C27B0', 'highlight': 'FFFACD'},
            'default': {'header': '1E90FF', 'accent': '87CEEB', 'highlight': 'FFD700'}
        }
    
    def detect_data_context(self, df: pd.DataFrame) -> str:
        """Détecte automatiquement le contexte des données"""
        text_content = ' '.join(df.astype(str).values.flatten().tolist()).lower()
        
        # Mots-clés pour chaque contexte
        contexts = {
            'geographic': ['latitude', 'longitude', 'coordonnées', 'gps', 'localisation', 
                          'adresse', 'ville', 'pays', 'région', 'carte', 'zone'],
            'financial': ['prix', 'montant', 'coût', 'budget', 'dépense', 'revenu',
                         'facture', 'paiement', 'transaction', 'devise'],
            'scientific': ['mesure', 'analyse', 'échantillon', 'expérience', 'résultat',
                          'donnée', 'valeur', 'statistique', 'métrique'],
            'events': ['invité', 'participant', 'table', 'vip', 'événement', 'date',
                      'horaire', 'lieu', 'organisateur']
        }
        
        scores = {}
        for context, keywords in contexts.items():
            score = sum(1 for kw in keywords if kw in text_content)
            scores[context] = score
        
        return max(scores, key=scores.get) if max(scores.values()) > 0 else 'default'
    
    def extract_coordinates(self, text: str) -> Optional[Dict[str, float]]:
        """Extrait les coordonnées GPS depuis du texte"""
        # Format: lat, lon ou (lat, lon) ou lat lon
        patterns = [
            r'([-+]?\d+\.\d+)[,\s]+([-+]?\d+\.\d+)',
            r'\(?([-+]?\d+\.\d+)[,\s]+([-+]?\d+\.\d+)\)?',
            r'lat[:\s]*([-+]?\d+\.\d+)[,\s]*lon[:\s]*([-+]?\d+\.\d+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, str(text), re.IGNORECASE)
            if match:
                return {'latitude': float(match.group(1)), 'longitude': float(match.group(2))}
        
        return None
    
    def classify_column(self, series: pd.Series) -> Dict[str, Any]:
        """Classifie automatiquement une colonne"""
        sample_text = ' '.join(series.astype(str).head(10).tolist())
        
        classification = {
            'name': series.name,
            'dtype': str(series.dtype),
            'type': 'unknown',
            'subtype': None,
            'stats': {},
            'patterns': []
        }
        
        # Détecter le type de données
        for data_type, pattern in self.data_types.items():
            if re.search(pattern, sample_text):
                classification['patterns'].append(data_type)
        
        # Classification par type pandas
        if pd.api.types.is_numeric_dtype(series):
            classification['type'] = 'numeric'
            classification['stats'] = {
                'min': series.min(),
                'max': series.max(),
                'mean': series.mean(),
                'median': series.median(),
                'std': series.std()
            }
            
            # Sous-types numériques
            if all(series.dropna().apply(lambda x: x == int(x))):
                classification['subtype'] = 'integer'
            elif series.min() >= -90 and series.max() <= 90:
                classification['subtype'] = 'potential_latitude'
            elif series.min() >= -180 and series.max() <= 180:
                classification['subtype'] = 'potential_longitude'
                
        elif pd.api.types.is_datetime64_any_dtype(series):
            classification['type'] = 'datetime'
            classification['stats'] = {
                'earliest': series.min(),
                'latest': series.max(),
                'range_days': (series.max() - series.min()).days if not series.isna().all() else 0
            }
            
        elif pd.api.types.is_string_dtype(series):
            classification['type'] = 'text'
            classification['stats'] = {
                'unique_values': series.nunique(),
                'most_common': series.mode()[0] if not series.empty else None,
                'avg_length': series.str.len().mean() if hasattr(series.str, 'len') else 0
            }
            
            # Sous-types textuels
            if 'coordinates' in classification['patterns']:
                classification['subtype'] = 'geographic_coordinates'
            elif 'email' in classification['patterns']:
                classification['subtype'] = 'email'
            elif 'phone' in classification['patterns']:
                classification['subtype'] = 'phone_number'
            elif 'url' in classification['patterns']:
                classification['subtype'] = 'url'
        
        return classification
    
    def auto_organize_dataframe(self, df: pd.DataFrame, context: str = None) -> pd.DataFrame:
        """Organise automatiquement un DataFrame selon son contexte"""
        if context is None:
            context = self.detect_data_context(df)
        
        # Classifications de toutes les colonnes
        classifications = {col: self.classify_column(df[col]) for col in df.columns}
        
        # Réorganiser les colonnes selon leur importance contextuelle
        priority_order = []
        
        if context == 'geographic':
            # Priorité: coordonnées, localisation, description
            priority = ['coordinates', 'latitude', 'longitude', 'location', 'address', 'city']
        elif context == 'financial':
            # Priorité: montants, dates, catégories
            priority = ['amount', 'price', 'cost', 'date', 'category', 'description']
        elif context == 'scientific':
            # Priorité: mesures, identifiants, paramètres
            priority = ['id', 'measurement', 'value', 'unit', 'timestamp', 'sample']
        elif context == 'events':
            # Priorité: rang, noms, tables, statuts
            priority = ['rank', 'rang', 'name', 'nom', 'table', 'status', 'vip']
        else:
            priority = []
        
        # Trier les colonnes
        sorted_cols = []
        for p in priority:
            for col in df.columns:
                if p.lower() in col.lower() and col not in sorted_cols:
                    sorted_cols.append(col)
        
        # Ajouter les colonnes restantes
        for col in df.columns:
            if col not in sorted_cols:
                sorted_cols.append(col)
        
        df_organized = df[sorted_cols].copy()
        
        # Traiter les colonnes spéciales
        for col in df_organized.columns:
            classif = classifications[col]
            
            # Extraire coordonnées si détectées
            if classif['subtype'] == 'geographic_coordinates':
                coords = df_organized[col].apply(self.extract_coordinates)
                if coords.notna().any():
                    df_organized[f'{col}_latitude'] = coords.apply(lambda x: x['latitude'] if x else None)
                    df_organized[f'{col}_longitude'] = coords.apply(lambda x: x['longitude'] if x else None)
            
            # Formater les dates
            if classif['type'] == 'datetime':
                df_organized[col] = pd.to_datetime(df_organized[col], errors='coerce')
            
            # Normaliser les textes
            if classif['type'] == 'text' and classif['stats'].get('avg_length', 0) < 100:
                df_organized[col] = df_organized[col].str.strip().str.title()
        
        return df_organized
    
    def create_intelligent_excel(self, 
                                  data: Dict[str, pd.DataFrame], 
                                  filename: str = None,
                                  add_charts: bool = True,
                                  add_conditional_formatting: bool = True) -> bytes:
        """
        Crée un fichier Excel intelligent avec mise en forme automatique
        
        Args:
            data: Dict {nom_onglet: DataFrame}
            filename: Nom du fichier (optionnel)
            add_charts: Ajouter des graphiques automatiques
            add_conditional_formatting: Ajouter formatage conditionnel
        
        Returns:
            bytes: Contenu du fichier Excel
        """
        output = BytesIO()
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        for sheet_name, df in data.items():
            # Organiser automatiquement
            context = self.detect_data_context(df)
            df_organized = self.auto_organize_dataframe(df, context)
            
            # Créer l'onglet
            ws = wb.create_sheet(title=str(sheet_name)[:31])
            
            # Obtenir le schéma de couleurs
            colors = self.color_schemes.get(context, self.color_schemes['default'])
            
            # Styles pour l'en-tête
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_fill = PatternFill(start_color=colors['header'], 
                                     end_color=colors['header'], 
                                     fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Border
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Écrire les en-têtes
            for col_idx, col_name in enumerate(df_organized.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Écrire les données
            for row_idx, row_data in enumerate(df_organized.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                    
                    # Formatage selon le type
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
                    elif isinstance(value, datetime):
                        cell.number_format = 'DD/MM/YYYY HH:MM'
            
            # Auto-ajuster la largeur des colonnes
            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Geler la première ligne
            ws.freeze_panes = 'A2'
            
            # Formatage conditionnel
            if add_conditional_formatting and len(df_organized) > 0:
                for col_idx, col_name in enumerate(df_organized.columns, 1):
                    classif = self.classify_column(df_organized[col_name])
                    
                    if classif['type'] == 'numeric':
                        # Échelle de couleurs pour les valeurs numériques
                        col_letter = ws.cell(row=1, column=col_idx).column_letter
                        ws.conditional_formatting.add(
                            f'{col_letter}2:{col_letter}{len(df_organized)+1}',
                            ColorScaleRule(
                                start_type='min', start_color='63BE7B',
                                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                end_type='max', end_color='F8696B'
                            )
                        )
            
            # Ajouter des graphiques automatiques
            if add_charts and len(df_organized) > 1:
                # Trouver les colonnes numériques pour graphiques
                numeric_cols = [col for col in df_organized.columns 
                               if pd.api.types.is_numeric_dtype(df_organized[col])]
                
                if numeric_cols:
                    # Graphique en barres pour la première colonne numérique
                    chart = BarChart()
                    chart.title = f"Analyse: {numeric_cols[0]}"
                    chart.style = 10
                    chart.height = 10
                    chart.width = 20
                    
                    data_ref = Reference(ws, min_col=df_organized.columns.get_loc(numeric_cols[0])+1,
                                        min_row=1, max_row=min(len(df_organized)+1, 50))
                    chart.add_data(data_ref, titles_from_data=True)
                    
                    # Placer le graphique à droite des données
                    chart_col = len(df_organized.columns) + 2
                    ws.add_chart(chart, f'{chr(65+chart_col)}2')
        
        wb.save(output)
        return output.getvalue()
    
    def analyze_and_structure(self, raw_data: Any, data_type: str = 'auto') -> Dict[str, pd.DataFrame]:
        """
        Analyse et structure automatiquement des données brutes
        
        Args:
            raw_data: Données brutes (texte, dict, list, DataFrame, etc.)
            data_type: Type de données ('auto', 'text', 'json', 'csv', 'coordinates')
        
        Returns:
            Dict de DataFrames organisés par catégorie
        """
        structured_data = {}
        
        if data_type == 'auto':
            # Détection automatique du type
            if isinstance(raw_data, pd.DataFrame):
                structured_data['Données principales'] = self.auto_organize_dataframe(raw_data)
            
            elif isinstance(raw_data, str):
                # Essayer de parser comme JSON
                try:
                    json_data = json.loads(raw_data)
                    if isinstance(json_data, list):
                        df = pd.DataFrame(json_data)
                        structured_data['Données JSON'] = self.auto_organize_dataframe(df)
                    elif isinstance(json_data, dict):
                        for key, value in json_data.items():
                            if isinstance(value, list):
                                df = pd.DataFrame(value)
                                structured_data[key] = self.auto_organize_dataframe(df)
                except:
                    # Parser comme texte/CSV
                    lines = [l.strip() for l in raw_data.split('\n') if l.strip()]
                    
                    # Détecter les coordonnées
                    coords_data = []
                    for line in lines:
                        coord = self.extract_coordinates(line)
                        if coord:
                            coords_data.append({
                                'source': line,
                                'latitude': coord['latitude'],
                                'longitude': coord['longitude']
                            })
                    
                    if coords_data:
                        structured_data['Coordonnées'] = pd.DataFrame(coords_data)
            
            elif isinstance(raw_data, list):
                if all(isinstance(item, dict) for item in raw_data):
                    df = pd.DataFrame(raw_data)
                    structured_data['Données liste'] = self.auto_organize_dataframe(df)
                else:
                    structured_data['Données'] = pd.DataFrame({'valeurs': raw_data})
            
            elif isinstance(raw_data, dict):
                for key, value in raw_data.items():
                    if isinstance(value, (list, pd.DataFrame)):
                        df = pd.DataFrame(value) if isinstance(value, list) else value
                        structured_data[key] = self.auto_organize_dataframe(df)
        
        return structured_data if structured_data else {'Données': pd.DataFrame([raw_data])}


# Fonction principale pour être appelée par l'IA
def organize_excel_with_ai(data: Any, 
                          filename: str = "organized_data.xlsx",
                          **options) -> Tuple[bytes, Dict[str, Any]]:
    """
    Fonction principale pour organiser des données Excel avec IA
    
    Args:
        data: Données à organiser (DataFrame, dict, list, text, etc.)
        filename: Nom du fichier de sortie
        **options: Options additionnelles (add_charts, add_conditional_formatting, etc.)
    
    Returns:
        Tuple[bytes, Dict]: (Contenu Excel, Rapport d'analyse)
    """
    organizer = ExcelAIOrganizer()
    
    # Structurer les données
    structured_data = organizer.analyze_and_structure(data)
    
    # Créer le rapport d'analyse
    report = {
        'timestamp': datetime.now().isoformat(),
        'filename': filename,
        'sheets': {},
        'total_rows': 0,
        'total_columns': 0
    }
    
    for sheet_name, df in structured_data.items():
        context = organizer.detect_data_context(df)
        classifications = {col: organizer.classify_column(df[col]) for col in df.columns}
        
        report['sheets'][sheet_name] = {
            'context': context,
            'rows': len(df),
            'columns': len(df.columns),
            'column_types': {col: classif['type'] for col, classif in classifications.items()},
            'has_coordinates': any(c['subtype'] == 'geographic_coordinates' for c in classifications.values())
        }
        
        report['total_rows'] += len(df)
        report['total_columns'] += len(df.columns)
    
    # Générer l'Excel
    excel_data = organizer.create_intelligent_excel(
        structured_data, 
        filename,
        add_charts=options.get('add_charts', True),
        add_conditional_formatting=options.get('add_conditional_formatting', True)
    )
    
    return excel_data, report


# ===============================================
# Wrapper BaseTool pour intégration dynamique
# ===============================================
try:
    import sys
    from pathlib import Path
    
    # Ajouter le dossier outils au path
    outils_dir = Path(__file__).parent
    if str(outils_dir) not in sys.path:
        sys.path.insert(0, str(outils_dir))
    
    from . import BaseTool
    
    class ExcelOrganizerTool(BaseTool):
        """Outil d'organisation intelligente Excel avec IA"""
        
        @property
        def name(self) -> str:
            return "excel_organizer"
        
        @property
        def description(self) -> str:
            return "Organise intelligemment des données dans Excel (coordonnées GPS, tableaux, classifications, etc.)"
        
        @property
        def capabilities(self) -> list:
            return [
                "Organisation automatique de données Excel",
                "Détection de types de données (GPS, téléphones, emails, dates)",
                "Structuration intelligente selon contexte",
                "Ajout de graphiques et mise en forme conditionnelle",
                "Support coordonnées géographiques",
                "Classification multi-critères"
            ]
        
        def can_handle(self, query: str, context: dict = None) -> float:
            """Détermine si l'outil peut traiter la requête"""
            query_lower = query.lower()
            
            # Mots-clés Excel
            excel_keywords = ['excel', 'xlsx', 'tableur', 'spreadsheet', 'feuille de calcul']
            # Mots-clés organisation
            org_keywords = ['organise', 'organiser', 'structure', 'trie', 'classe', 'range']
            # Mots-clés données
            data_keywords = ['données', 'data', 'coordonnées', 'gps', 'tableau', 'liste']
            
            score = 0.0
            
            # Score principal: mention Excel
            if any(kw in query_lower for kw in excel_keywords):
                score += 0.4
            
            # Score organisation
            if any(kw in query_lower for kw in org_keywords):
                score += 0.3
            
            # Score données
            if any(kw in query_lower for kw in data_keywords):
                score += 0.2
            
            # Bonus si fichier Excel dans contexte
            if context and context.get('file_type') in ['.xlsx', '.xls', '.csv']:
                score += 0.3
            
            return min(score, 1.0)
        
        def execute(self, query: str, context: dict = None) -> dict:
            """Exécute l'organisation Excel"""
            try:
                # Récupérer les données depuis le contexte
                if not context or 'data' not in context:
                    return {
                        'success': False,
                        'error': 'Aucune donnée fournie. Uploadez un fichier Excel ou CSV.'
                    }
                
                data = context['data']
                filename = context.get('filename', 'organized_data.xlsx')
                options = context.get('options', {})
                
                # Organiser les données
                excel_bytes, report = organize_excel_data(data, filename, options)
                
                return {
                    'success': True,
                    'excel_data': excel_bytes,
                    'filename': filename,
                    'report': report,
                    'message': f"✅ Excel organisé avec succès: {report['total_rows']} lignes, {report['total_columns']} colonnes"
                }
                
            except Exception as e:
                return {
                    'success': False,
                    'error': f"Erreur lors de l'organisation Excel: {str(e)}"
                }
    
    # Instance exportable
    excel_organizer_tool = ExcelOrganizerTool()
    
except ImportError:
    # BaseTool non disponible, mode standalone
    excel_organizer_tool = None
    print("⚠️ ExcelOrganizerTool: BaseTool non disponible, mode standalone uniquement")
